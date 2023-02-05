import boto3
import json
from openpyxl import Workbook

all_ports = ['Ramsgate', 'Cowes', 'Plymouth', 'Milford Haven', 'Dublin', 'Oban', 'Arbroath', 'Blyth', 'OGA60, Suffolk Yacht Harbour, River Orwell']
party_name_map = { 
    'OGA60, Suffolk Yacht Harbour, River Orwell': 'Jubilee Party',
    'Milford Haven': 'Neyland',
    'Arbroath': 'River Tay',

}
def map_ports(ports):
    p = {}
    for port in all_ports: 
        if port in ports:
            p[port] = '✓'
        else:
            p[port] = ''
    return p

def skipper(user, payer):
    if 'name' in user:
        return user['name']
    if 'name' in payer:
        name = payer['name']
        return f"{name['surname']} {name['given_name']}"
    return 'not known'

def boat_entry(d, email):
    b = { 'boat': d['boat']['name'], 'oga_no': int(d['boat']['oga_no']) }
    user = d.get('user', {})
    if 'payment' in d:
        payer = d['payment']['payer']
    else:
        payer = {}
    n = {}
    n['skipper'] = skipper(user, payer)
    n['email'] = email
    return {**b, **n}

def entry(item):
    d = item['data']
    ports = d.get('port', {})
    return {**boat_entry(d, item['email']), **map_ports(ports) }

def crewing(items):
    legs = {}
    for port in all_ports:
        legs[port] = []
    for item in items:
        d = item['data']
        boat = d['boat']
        offers = d.get('leg', [])
        if len(offers) > 0:
            for offer in offers:
                # print(offer)
                o = { 'name': boat['name'], 'oga_no': int(boat['oga_no']), 'email': item['email'], 'spaces': int(offer['spaces'])}
                # print(o)
                legs[offer['from']].append(o)

ddb = boto3.resource('dynamodb')
table = ddb.Table('member_entries')
sf = {'topic': { 'AttributeValueList': ['RBC 60'], 'ComparisonOperator': 'EQ'}}
t = table.scan(ScanFilter=sf)
data = []
for item in t['Items']:
    data.append((entry(item)))
# print(data)

workbook = Workbook()
ws=workbook.active
ws.title = 'Entries'
ws.append(list(data[0].keys()))
for row in data:
    ws.append(list(row.values()))
public_table = ddb.Table('public_fleets')
for port in all_ports:
    boats = []
    port_name = party_name_map.get(port, port)
    port_sheet = workbook.create_sheet(port_name)
    columns = ['boat','oga_no','skipper','email']
    port_sheet.append(columns)
    for boat in data:
        if boat[port]:
            boats.append(boat['oga_no'])
            b = [boat[k] for k in columns]
            port_sheet.append(b)
    public_table.put_item(Item={
        'name': f"RBC 60 {port_name}",
        'owner_gold_id': 559,
        'filters': {
            'oga_nos': boats
        },
        'public': True,
    })

# crewing_sheet = workbook.create_sheet("Crewing")
# crew = crewing(t['Items'])
workbook.save("rbc 60.xlsx")

